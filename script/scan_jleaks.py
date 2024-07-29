from pathlib import Path
import time
import re
import logging
import random
import traceback
from tqdm import tqdm
from openpyxl import load_workbook
from srctoolkit import javalang
from srctoolkit.dependency_analyzer import DependencyAnalyzer

from app.llms import LLM_FACTORY
from app.prompts import PROMPT_FACTORY
from app.llm4leak import LLM4Leak
from app.logging_util import init_log_file

ANALYZER = DependencyAnalyzer()

if __name__ == "__main__":
    BUGGY_DIR = "data/JLeaks/JLeaksDataset/bug_method"
    FIXED_DIR = "data/JLeaks/JLeaksDataset/fix_method"
    META_PATH = "data/JLeaks/JLeaksDataset/JLeaks.xlsx"

    # MODEL = "gpt-4"
    # MODEL = "gpt-4-turbo"
    # MODEL = "gpt-3.5-turbo"
    # MODEL = "llama-3-8b"
    MODEL = "gemma-2-9b"

    PROMPT = "inferroi-paper"

    LOG_FILE = f"log/jleaks-{MODEL}-{PROMPT}.log"

    detector = LLM4Leak(llm=LLM_FACTORY[MODEL](), prompt_template=PROMPT_FACTORY[PROMPT])

    init_log_file(LOG_FILE)

    leaked_type_dict = dict()
    all_resource_types = set()
    workbook = load_workbook(filename=META_PATH)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        _id, _type1, _type2 = row[0], row[23], row[24]
        if _type1:
            leaked_type_dict[_id] = {t.split(".")[-1] for t in _type1.split(",")}
            all_resource_types.update(leaked_type_dict[_id])
        if _type2:
            leaked_type_dict[_id] = {t.split(".")[-1] for t in _type2.split(",")}
            all_resource_types.update(leaked_type_dict[_id])
    print(all_resource_types)

    exceptional_ids = {741} 

    tp, fp, fn = 0, 0, 0
    insts = []
    for buggy_path, fixed_path in zip(sorted(list(Path(BUGGY_DIR).glob("*.java"))), sorted(list(Path(FIXED_DIR).glob("*.java")))):
        buggy_id = int(buggy_path.parts[-1].split("-")[1])
        fixed_id = int(fixed_path.parts[-1].split("-")[1])
        if buggy_id != fixed_id:
            continue
        if buggy_id in exceptional_ids:
            continue
        if buggy_id not in leaked_type_dict:
            continue
        with buggy_path.open("r") as bf, fixed_path.open("r") as ff:
            buggy_method, fixed_method = bf.read(), ff.read()
            fixed_method = buggy_method.split("{")[0].lstrip() + fixed_method
        if re.sub(r"\s+", "", buggy_method, flags=re.S|re.M) == re.sub(r"\s+", "", fixed_method, flags=re.S|re.M):
            continue
        if len(buggy_method.split("\n")) > 100 or len(fixed_method.split("\n")) > 100:
            continue
        try:
            javalang.parse.parse('class Foo{\n' + buggy_method + '\n}')
            ANALYZER.build_cfg('class Foo{\n' + buggy_method + '\n}')
        except Exception:
            # traceback.print_exc()
            continue
        try:
            javalang.parse.parse('class Foo{\n' + fixed_method + '\n}')
            ANALYZER.build_cfg('class Foo{\n' + fixed_method + '\n}')
        except Exception:
            # traceback.print_exc()
            continue
        insts.append((buggy_id, buggy_path, fixed_path))

    random.shuffle(insts)

    for ID, buggy_path, fixed_path in tqdm(insts):
        with buggy_path.open("r") as bf, fixed_path.open("r") as ff:
            buggy_method, fixed_method = bf.read(), ff.read()
            fixed_method = buggy_method.split("{")[0].lstrip() + fixed_method
        resource_types = leaked_type_dict[ID]
        logging.info(f"########### ID-{ID} ###########")
        logging.info(f"resource type: {resource_types}")

        logging.info("###### BUGGY METHOD ######")
        logging.info(buggy_path)
        logging.info(buggy_method)
        _, _, _, leaks = detector.detect(buggy_method, consider_exception=True)
        time.sleep(5)
        leaked_types = {leak["type"].split(".")[-1] for leak in leaks.values()}
        normalized_leaked_types = set()
        for leaked_type in leaked_types:
            leaked_type:str
            if leaked_type[0].islower() and leaked_type not in resource_types:
                for resource_type in resource_types:
                    if resource_type.lower() in leaked_type.lower():
                        resource_type = resource_type
            normalized_leaked_types.add(leaked_type)
        leaked_types = normalized_leaked_types
        logging.info(resource_types & leaked_types)
        if len(resource_types & leaked_types) > 0:
            tp += 1
        else:
            fn += 1
        logging.info("###### FIXED METHOD ######")
        logging.info(fixed_path)
        logging.info(fixed_method)
        _, _, _, leaks = detector.detect(fixed_method, consider_exception=True)
        time.sleep(5)
        leaked_types = {leak["type"].split(".")[-1] for leak in leaks.values()}
        normalized_leaked_types = set()
        for leaked_type in leaked_types:
            leaked_type:str
            if leaked_type[0].islower() and leaked_type not in resource_types:
                for resource_type in resource_types:
                    if resource_type.lower() in leaked_type.lower():
                        resource_type = resource_type
            normalized_leaked_types.add(leaked_type)
        leaked_types = normalized_leaked_types
        logging.info(resource_types & leaked_types)
        if len(resource_types & leaked_types) > 0:
            fp += 1

        logging.info(f"tp: {tp}, fp: {fp}, fn: {fn}")
        logging.info(f"precision: {(tp / (tp + fp)) if tp > 0 else 0.}, recall: {(tp / (tp + fn)) if tp > 0 else 0.}")


    identified_types = set()
    with Path(LOG_FILE).open("r") as f:
        for mobj in re.finditer(r"intentions: \[(.*?)\]", f.read()):
            intentions = mobj.group(1)
            for mobj2 in re.finditer(r"\(\d+, '\w+', '\w+', '(\w+)'\)", intentions):
                _type = mobj2.group(1).split(".")[-1]
                # if _type[0].islower():
                #     continue
                identified_types.add(_type)
    logging.info(identified_types)
    
    logging.info(f"{len(all_resource_types)}, {len(identified_types)}, {len(identified_types & all_resource_types)}")
    logging.info(len(identified_types & all_resource_types) / len(all_resource_types))